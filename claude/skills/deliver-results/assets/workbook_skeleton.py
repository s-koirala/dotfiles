"""xlsxwriter workbook skeleton with the canonical 7-sheet layout.

Sheets emitted in fixed order:
    README -> parameters -> methods -> results_* -> figures -> audit_trail -> references

README sheet header carries the 13-field ReproLog envelope (R1-A) so a future
auditor can reproduce the analysis from the workbook alone (paired with
git HEAD and the dataset manifest).

xlsxwriter rationale (per memo §B.5):
- write-only library (no round-trip; if you need to edit a user-supplied
  workbook, fall back to openpyxl ad-hoc)
- native chart object model (better than openpyxl for embedding)
- constant-memory mode for >1M rows (https://xlsxwriter.readthedocs.io/working_with_memory.html)
- sparklines + broader conditional-formatting rule set

Usage:
    from workbook_skeleton import build_workbook
    build_workbook(
        out_path="artifacts/workbooks/results_2026-05-15.xlsx",
        title="Hypothesis H001 v0 — KPI report card",
        repro_log_path="logs/reproducibility/repro_log_<id>.json",
        results_blocks=[
            ("results_oos", oos_df),
            ("results_walkforward", wf_df),
        ],
    )

Self-test: `python workbook_skeleton.py --selftest` writes a fixture workbook
to a tempdir, verifies all 7 sheets exist + README header maps to ReproLog
schema keys, exits 0.
"""
from __future__ import annotations

import json
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Sequence

# Canonical sheet order (fixed; do not customize). README is FIRST so a
# reader opening the workbook lands on metadata; audit_trail is LAST so it
# accumulates without disturbing the analytical content.
_SHEET_ORDER = (
    "README",
    "parameters",
    "methods",
    "results",   # may be split into multiple results_* sheets
    "figures",
    "audit_trail",
    "references",
)

# Map ReproLog schema keys to README sheet row labels.
_REPRO_LOG_FIELDS = (
    "run_id", "phase", "hypothesis_id", "timestamp_utc", "git_head",
    "pip_freeze_sha256", "pip_freeze_path", "dataset_checksums",
    "rng_seed", "model_hash", "config_resolved_sha256", "host", "env_id",
)


def build_workbook(
    out_path: Path | str,
    title: str,
    repro_log_path: Path | str | None = None,
    results_blocks: Sequence[tuple[str, "pd.DataFrame"]] | None = None,
    methods_md: str = "",
    parameters: dict[str, object] | None = None,
    references: Sequence[dict] | None = None,
) -> Path:
    """Build the canonical workbook. Returns the output Path.

    Args:
        out_path: workbook destination (.xlsx).
        title: human-readable workbook title for README sheet.
        repro_log_path: path to a ReproLog JSON (R1-A schema) to inline into README.
        results_blocks: list of (sheet_name, DataFrame) pairs. Each emits its
                        own sheet named `results_<sheet_name>`. If None or
                        empty, a single empty `results` placeholder sheet is created.
        methods_md: markdown text for the methods sheet (literal cell content).
        parameters: dict of run-time parameters; one row per key.
        references: list of dicts with fields {type, authors, title, year, doi}.
    """
    try:
        import xlsxwriter
    except ImportError:
        raise RuntimeError(
            "xlsxwriter is required. Install with `uv pip install xlsxwriter`."
        )

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = xlsxwriter.Workbook(str(out_path))

    # Format objects
    fmt_h1 = wb.add_format({"bold": True, "font_size": 14})
    fmt_h2 = wb.add_format({"bold": True, "font_size": 11})
    fmt_label = wb.add_format({"bold": True, "valign": "top"})
    fmt_code = wb.add_format({"font_name": "Consolas", "valign": "top"})

    # --- README sheet ---
    s = wb.add_worksheet("README")
    s.set_column(0, 0, 28)
    s.set_column(1, 1, 60)
    s.write(0, 0, title, fmt_h1)
    s.write(1, 0, f"Generated UTC: {datetime.now(timezone.utc).isoformat(timespec='seconds')}")
    s.write(3, 0, "ReproLog envelope (R1-A schema)", fmt_h2)
    repro_data = _load_repro_log(repro_log_path)
    for i, field in enumerate(_REPRO_LOG_FIELDS, start=4):
        s.write(i, 0, field, fmt_label)
        value = repro_data.get(field, "<missing>")
        if isinstance(value, (dict, list)):
            s.write(i, 1, json.dumps(value, sort_keys=True, ensure_ascii=False), fmt_code)
        else:
            s.write(i, 1, str(value), fmt_code)

    # --- parameters sheet ---
    s = wb.add_worksheet("parameters")
    s.set_column(0, 0, 32)
    s.set_column(1, 1, 40)
    s.write(0, 0, "Parameter", fmt_label)
    s.write(0, 1, "Value", fmt_label)
    for i, (k, v) in enumerate(sorted((parameters or {}).items()), start=1):
        s.write(i, 0, k)
        s.write(i, 1, str(v))

    # --- methods sheet ---
    s = wb.add_worksheet("methods")
    s.set_column(0, 0, 100)
    if methods_md:
        for i, line in enumerate(methods_md.splitlines()):
            s.write(i, 0, line, fmt_code)
    else:
        s.write(0, 0, "<<TODO: methodology summary; cite the analysis design.md per R3-2a>>", fmt_label)

    # --- results sheets (one per block; or a placeholder) ---
    if results_blocks:
        for sheet_name, df in results_blocks:
            full_name = f"results_{sheet_name}"[:31]  # Excel 31-char sheet name limit
            ws = wb.add_worksheet(full_name)
            _write_dataframe(ws, df, fmt_label)
    else:
        s = wb.add_worksheet("results")
        s.write(0, 0, "<<TODO: results>>", fmt_label)

    # --- figures sheet ---
    s = wb.add_worksheet("figures")
    s.write(0, 0, "Embed PNG@300dpi figures via insert_image. "
                  "Use save_figure(target='single_col') from R2-C.", fmt_label)

    # --- audit_trail sheet ---
    s = wb.add_worksheet("audit_trail")
    s.set_column(0, 0, 14)
    s.set_column(1, 1, 12)
    s.set_column(2, 2, 60)
    s.write(0, 0, "Round", fmt_label)
    s.write(0, 1, "Severity", fmt_label)
    s.write(0, 2, "Issue / disposition", fmt_label)
    # Populated by audit-remediate-loop post-emission

    # --- references sheet ---
    s = wb.add_worksheet("references")
    s.set_column(0, 0, 8)
    s.set_column(1, 1, 80)
    s.write(0, 0, "Cite#", fmt_label)
    s.write(0, 1, "Reference", fmt_label)
    for i, ref in enumerate(references or [], start=1):
        s.write(i, 0, f"[{i}]")
        # Format: "Authors (Year). Title. Journal. DOI."
        authors = ref.get("authors", "")
        s.write(i, 1, f"{authors} ({ref.get('year','?')}). "
                      f"{ref.get('title','')}. {ref.get('journal','')}. "
                      f"DOI {ref.get('doi','')}")

    wb.close()
    return out_path


def _load_repro_log(p: Path | str | None) -> dict:
    if p is None:
        return {f: "<not provided>" for f in _REPRO_LOG_FIELDS}
    try:
        return json.loads(Path(p).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        return {"_error": str(e)}


def _write_dataframe(ws, df, fmt_label) -> None:
    """Minimal DataFrame writer (no pandas import here; pandas is optional)."""
    if df is None:
        ws.write(0, 0, "<no data>", fmt_label)
        return
    # Write header
    cols = list(df.columns) if hasattr(df, "columns") else []
    for j, col in enumerate(cols):
        ws.write(0, j, str(col), fmt_label)
    # Write rows
    values = df.values if hasattr(df, "values") else []
    for i, row in enumerate(values, start=1):
        for j, val in enumerate(row):
            try:
                ws.write(i, j, val)
            except TypeError:
                ws.write(i, j, str(val))


def _selftest() -> int:
    """Self-test: fixture workbook with no pandas dependency."""
    try:
        import xlsxwriter  # noqa: F401
    except ImportError:
        print("FAIL: xlsxwriter not installed; install with `uv pip install xlsxwriter`",
              file=sys.stderr)
        return 1

    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "fixture.xlsx"
        # No results blocks, no pandas — exercise the empty-results path
        build_workbook(
            out_path=out,
            title="Fixture",
            repro_log_path=None,
            results_blocks=None,
            methods_md="Fixture methodology.",
            parameters={"alpha": 0.05, "seed": 42},
            references=[
                {"authors": "Wong B", "year": 2011, "title": "Color blindness",
                 "journal": "Nat Methods", "doi": "10.1038/nmeth.1618"},
            ],
        )

        # Reopen with openpyxl to verify sheet inventory (if available)
        try:
            from openpyxl import load_workbook
        except ImportError:
            print(f"PASS (limited): xlsxwriter produced {out} "
                  f"({out.stat().st_size} bytes); openpyxl not installed so "
                  f"sheet-inventory verification skipped.")
            return 0

        wb = load_workbook(out, read_only=True)
        expected_sheets = {"README", "parameters", "methods", "results",
                           "figures", "audit_trail", "references"}
        got = set(wb.sheetnames)
        missing = expected_sheets - got
        if missing:
            print(f"FAIL: missing sheets: {missing}", file=sys.stderr)
            return 2

        # Verify README has all 13 ReproLog field labels
        rd = wb["README"]
        cells = [(row[0].value, row[1].value) for row in rd.iter_rows(min_row=5, max_row=17, max_col=2)]
        readme_labels = {label for (label, _) in cells if label}
        expected_fields = set(_REPRO_LOG_FIELDS)
        missing_fields = expected_fields - readme_labels
        if missing_fields:
            print(f"FAIL: README missing ReproLog fields: {missing_fields}",
                  file=sys.stderr)
            return 3

        print(f"PASS: xlsxwriter produced {out} ({out.stat().st_size} bytes); "
              f"7 sheets present; README has all 13 ReproLog field labels.")
        return 0


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        sys.exit(_selftest())
    print("Usage: python workbook_skeleton.py --selftest", file=sys.stderr)
    sys.exit(64)
